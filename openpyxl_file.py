#tutorial from http://zetcode.com/articles/openpyxl/
# styles for cells are here http://openpyxl.readthedocs.io/en/default/styles.html

from openpyxl import Workbook
import time


wb = Workbook()
ws = wb.active
#does not work for me
#ws.sheet_properties.tabColor = "1072BA"
ws.title = 'Custom Title'

ws['A1'] = 56
ws['A2'] = 43


now = time.strftime("%x")
ws['A3'] = now

wb.save('C:/Users/Acer/Desktop/python files Ive done/sample.xlsx')
